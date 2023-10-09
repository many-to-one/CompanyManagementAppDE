import os
from django.conf import settings
from ..models import Vacations, Work
from django.db.models import Sum
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


####################################################################################
#                                       EXCEL                                      #
####################################################################################

def vacationsToExcel(request):
    user = request.session.get('user')
    year = request.session.get('year') 
    if user == '' and year is not None:
        vacations = Vacations.objects.filter(
            v_from__startswith=year,
        ).order_by('-id')
    elif user == '' and year == '':
        vacations = Vacations.objects.all().order_by('-id')
    elif user is not None and year is not None:
        vacations = Vacations.objects.filter(
            v_from__startswith=year,
            username=user
        ).order_by('-id')

    wb = openpyxl.Workbook()

    # Get the default active sheet
    sheet = wb.active

    # Add the table headers
    headers = [
        'UŻYTKOWNIK', 
        'DATA WNIOSKU', 
        'TYP',
        'OD',
        'DO',
        'ILOŚĆ DNI',
        'AKCEPTACJA',
        ]  # Specify your desired headers
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    # Add the table data
    for row_num, vacation in enumerate(vacations, 2):
        sheet.cell(row=row_num, column=1).value = vacation.username
        sheet.cell(row=row_num, column=2).value = vacation.date
        sheet.cell(row=row_num, column=3).value = vacation.type
        sheet.cell(row=row_num, column=4).value = vacation.v_from
        sheet.cell(row=row_num, column=5).value = vacation.v_to
        sheet.cell(row=row_num, column=6).value = vacation.days_planned
        if vacation.accepted == True:
            sheet.cell(row=row_num, column=7).value = 'Tak'
        else:
            sheet.cell(row=row_num, column=7).value = 'Nie'

    # Set the width of columns
    column_widths = [15, 15, 25, 15, 15, 10, 10]  # Specify the desired width for each column
    for col_num, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = width

    # Create a response to serve the Excel file for download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=urlopy.xlsx'
    wb.save(response)

    return response


def raportsToExcel(request):

    total_fields = {
                'total_coffee_food': 'coffee_food',
                'total_fuel': 'fuel',
                'total_prepayment': 'prepayment',
                'total_phone_costs': 'phone_costs',
                'total_payment': 'payment',
                'total_sum_time_sec': 'sum_time_sec',
                'total_sum_over_time_sec': 'sum_over_time_sec'
            }
    works = request.session.get('works')
    ids = [w['id'] for w in works]

    if request.method == 'POST' or works:
        visible_values = request.POST.getlist('visible_values[]')
        raports = Work.objects.filter(id__in=visible_values)
        if works:
            raports = Work.objects.filter(id__in=ids)

        total_coffee_food = float(raports.aggregate(total_coffee_food=Sum('coffee_food'))['total_coffee_food'])
        total_fuel = float(raports.aggregate(total_fuel=Sum('fuel'))['total_fuel'])
        total_prepayment = float(raports.aggregate(total_prepayment=Sum('prepayment'))['total_prepayment'])
        total_phone_costs = float(raports.aggregate(total_phone_costs=Sum('phone_costs'))['total_phone_costs'])
        total_payment = float(raports.aggregate(total_payment=Sum('payment'))['total_payment'])
        total_sum_time_sec = float(raports.aggregate(total_sum_time_sec=Sum('sum_time_sec'))['total_sum_time_sec'])
        total_sum_over_time_sec = float(raports.aggregate(total_sum_over_time_sec=Sum('sum_over_time_sec'))['total_sum_over_time_sec'])

        if total_sum_time_sec:
            ### total_sum_time_sec => hours:minutes ###
            total_hours = total_sum_time_sec // 3600
            total_sec = total_sum_time_sec % 3600
            total_min = total_sec // 60
            total_work_time = f'{int(total_hours)}:{int(total_min)}'
        else:
            total_work_time = '0:00'
        if total_sum_over_time_sec:
            ### total_sum_over_time_sec => hours:minutes ###
            total_hours = total_sum_over_time_sec // 3600
            total_sec = total_sum_over_time_sec % 3600
            total_min = total_sec // 60
            if total_min < 10 or total_min == 0.0:
                total_work_over_time = f'{int(total_hours)}:0{int(total_min)}'
            else:
                total_work_over_time = f'{int(total_hours)}:{int(total_min)}'
        else:
            total_work_over_time = '0:00'

        total = total_coffee_food + total_fuel + total_phone_costs + total_payment

        wb = openpyxl.Workbook()

        # Get the default active sheet
        sheet = wb.active

        # Add the table headers
        headers = [
            'Data', 
            'Pracownik', 
            'Objekt',
            'Początek',
            'Koniec',
            'Czas pracy',
            'Nadgodziny',
            'Czynność',
            'Kawa/Posiłki',
            'Zaliczka',
            'Paliwo',
            'Telefon',
            'Opłata',
            ]  # Specify your desired headers
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

        # Add the table data
        for row_num, raport in enumerate(raports, 3):
            sheet.cell(row=row_num, column=1).value = raport.date
            sheet.cell(row=row_num, column=2).value = raport.username
            sheet.cell(row=row_num, column=3).value = raport.work_object
            sheet.cell(row=row_num, column=4).value = raport.timestart
            sheet.cell(row=row_num, column=5).value = raport.timefinish
            sheet.cell(row=row_num, column=6).value = raport.diff_time
            sheet.cell(row=row_num, column=7).value = raport.over_time
            sheet.cell(row=row_num, column=8).value = raport.work_type
            sheet.cell(row=row_num, column=9).value = raport.coffee_food
            sheet.cell(row=row_num, column=10).value = raport.prepayment
            sheet.cell(row=row_num, column=11).value = raport.fuel
            sheet.cell(row=row_num, column=12).value = raport.phone_costs
            sheet.cell(row=row_num, column=13).value = raport.payment

        for raport in range(0, 16):
            column_values = ['RAZEM:', total, '', '', '', total_work_time, total_work_over_time, '', 
                            total_coffee_food, total_prepayment, total_fuel, 
                            total_phone_costs, total_payment, '']

        for column, value in enumerate(column_values, 1):
            sheet.cell(row=len(raports)+4, column=column).value = value

        # Set the width of columns
        column_widths = [15, 15, 40, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15]  # Specify the desired width for each column
        for col_num, width in enumerate(column_widths, 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = width

        # Specify the file name for the Excel file
        file_name = 'raporty.xlsx'

        # Construct the full file path
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)

        # Save the workbook to the file path
        wb.save(file_path)

        # Create a response to serve the Excel file for download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=raporty.xlsx'
        wb.save(response)

        return response
    

####################################################################################
#                                  END EXCEL                                       #
####################################################################################